from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from sqlalchemy import create_engine, text
import pandas as pd
import os
import re
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
app.secret_key = "your_secret_key"

# Configuración de la base de datos
DATABASE_URI = "mssql+pyodbc://sa:infinity@192.168.201.12/HERMESV5_HISTORICO?driver=ODBC+Driver+17+for+SQL+Server"

engine = create_engine(DATABASE_URI)

# Configurar el logging
logging.basicConfig(filename='activity.log', level=logging.INFO)

# Crear la carpeta 'uploads' si no existe
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def limpiar_telefono(telefono):
    if pd.isna(telefono):
        return None
    telefono = str(telefono)
    telefono = re.sub(r'[ .\-\/]', '', telefono)
    telefono = re.sub(r'^(\+34|0034)', '', telefono)
    if len(telefono) == 9 and telefono.isdigit():
        return telefono
    else:
        return None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        flash('No file part')
        return redirect(url_for('index'))

    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        return redirect(url_for('index'))

    if file:
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)

        # Procesar el archivo Excel
        df = pd.read_excel(filepath)

        # Renombrar las columnas a un formato estándar
        df.rename(columns=lambda x: x.strip().lower(), inplace=True)
        df.rename(columns={
            'telefono': 'telefono',
            'teléfono': 'telefono',
            'telf1': 'telefono',
            'tel1': 'telefono',
            'telefono1': 'telefono',
            'TELEFONO1': 'telefono',
            'TELEFONO': 'telefono',
            'phone': 'telefono'
        }, inplace=True)

        # Verificar si 'telefono' existe
        if 'telefono' not in df.columns:
            raise KeyError("No se encontró la columna 'telefono' en el DataFrame.")

        # Limpiar los teléfonos
        df['telefono_limpio'] = df['telefono'].apply(limpiar_telefono)

        # Conectar a la base de datos y ejecutar la lógica de scoring, media de intentos y media de intentos para positivo
        with engine.connect() as conn:
            for index, row in df.iterrows():
                telefono = str(row['telefono'])  # Convertir el número de teléfono a string

                # Consulta de scoring
                query_scoring = text("""
                    SELECT CASE 
                        WHEN COUNT(*) >= 20 THEN (CAST(SUM(CASE WHEN CallStatusNum < 11 THEN 1 ELSE 0 END) AS FLOAT) / COUNT(*)) * 100 
                        ELSE (SUM(CASE WHEN CallStatusNum < 11 THEN 1 ELSE 0 END) / 20.0) * 100 
                    END AS conteo 
                    FROM ODCalls WHERE ANI = :telefono
                """)
                result_scoring = conn.execute(query_scoring, {'telefono': telefono}).fetchone()
                conteo = result_scoring[0] if result_scoring else 0

                # Agregar logging para el resultado de scoring
                logging.info(f'Consulta para scoring: {query_scoring}, Parámetros: {telefono}, Resultado: {conteo}')
                df.at[index, 'scoring'] = conteo

                # Consulta de media de intentos para contacto
                query_media_intentos = text("""
                    SELECT 
                        AVG(intento) AS media_llamadas_para_contacto
                    FROM (
                        SELECT 
                            ROW_NUMBER() OVER (PARTITION BY ANI ORDER BY CallLocalTimeString) AS intento
                        FROM ODCalls
                        WHERE ANI = :telefono 
                            AND Duration > 1  -- Solo llamadas contestadas
                    ) AS llamadas
                """)
                result_media_intentos = conn.execute(query_media_intentos, {'telefono': telefono}).fetchone()
                media_intentos = result_media_intentos[0] if result_media_intentos else None
                df.at[index, 'media_intentos_para_contacto'] = media_intentos

                # Consulta de media de intentos para contacto positivo
                query_media_intentos_positivo = text("""
                    SELECT 
                        CASE 
                            WHEN COUNT(CASE WHEN Duration > 1 AND CallStatusNum BETWEEN 1 AND 10 THEN 1 END) = 0 THEN 0
                            ELSE COUNT(*) * 1.0 / COUNT(CASE WHEN Duration > 1 AND CallStatusNum BETWEEN 1 AND 10 THEN 1 END)
                        END AS media_llamadas_para_contacto_positivo
                    FROM ODCalls
                    WHERE ANI = :telefono
                    AND (Duration > 1 OR CallStatusNum BETWEEN 1 AND 10)
                """)
                result_media_intentos_positivo = conn.execute(query_media_intentos_positivo, {'telefono': telefono}).fetchone()

                # Agregar un log para depuración
                if result_media_intentos_positivo is None:
                    logging.info(f'No se encontraron resultados para el teléfono: {telefono}')
                else:
                    logging.info(f'Resultado para el teléfono {telefono}: {result_media_intentos_positivo[0]}')

                media_intentos_positivo = result_media_intentos_positivo[0] if result_media_intentos_positivo else None
                df.at[index, 'media_intentos_para_contacto_positivo'] = media_intentos_positivo

        # Ordenar el DataFrame de mayor a menor según el scoring
        df.sort_values(by='scoring', ascending=False, inplace=True)

        # Convertir los valores numéricos pequeños a un formato más legible con 2 decimales
        df['media_intentos_para_contacto_positivo'] = df['media_intentos_para_contacto_positivo'].apply(lambda x: '{:.2f}'.format(x) if pd.notna(x) else x)

        # Guardar el archivo Excel con el scoring, media de intentos y todos los datos
        output_filepath = os.path.join(UPLOAD_FOLDER, 'resultados_Scoring.xlsx')
        df.to_excel(output_filepath, index=False)

        # Aplicar estilo de colores en la columna 'scoring' usando openpyxl
        wb = load_workbook(output_filepath)
        ws = wb.active
        scoring_column_index = ws.max_column - 2  # Columna 'scoring'
        media_intentos_column_index = ws.max_column - 1  # Columna 'media_intentos_para_contacto'
        media_intentos_positivo_column_index = ws.max_column  # Última columna es 'media_intentos_para_contacto_positivo'

        for row in range(2, ws.max_row + 1):  # Empezamos en la fila 2 para evitar el encabezado
            score = ws.cell(row=row, column=scoring_column_index).value
            if score is not None:
                # Definir el color basado en el score (de verde a rojo)
                red = int(min(255, (255 - int(2.55 * score))))
                green = int(min(255, int(2.55 * score)))
                fill = PatternFill(start_color=f"{red:02X}{green:02X}00", end_color=f"{red:02X}{green:02X}00", fill_type="solid")
                ws.cell(row=row, column=scoring_column_index).fill = fill

        wb.save(output_filepath)
        wb.close()

        # Generar la tabla HTML para mostrar en el navegador
        resultados_limpieza = df.to_html(classes=['table', 'scoring'], justify='center')

        return render_template('resultados.html', tabla=resultados_limpieza)

@app.route('/download')
def download_file():
    # Descargar el archivo de resultados
    output_filepath = os.path.join(UPLOAD_FOLDER, 'resultados_Scoring.xlsx')
    return send_file(output_filepath, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5000)
