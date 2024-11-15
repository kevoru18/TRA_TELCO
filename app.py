from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from sqlalchemy import create_engine, text
import pandas as pd
import os
import re
import logging
import getpass  # Importar el módulo para obtener el nombre del usuario
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill 
from apscheduler.schedulers.background import BackgroundScheduler
import smtplib
from email.mime.text import MIMEText
from db_utils import obtener_datos_kpi, obtener_datos_clicks
from excel_utils import generar_graficos_excel


app = Flask(__name__)
app.secret_key = "your_secret_key"

# Ruta para la página de inicio
@app.route('/')
def inicio():
    return render_template('inicio.html')

@app.route('/index')
def index():
    return render_template('index.html')


# Ruta para la aplicación principal
@app.route('/main_app')
def main_app():
    return redirect(url_for('index'))  # Redirige a la función de carga de archivos

# Ruta para la página de gráficos
@app.route('/graficos', methods=['GET', 'POST'])
def graficos():
    if request.method == 'POST':
        db_name = request.form.get('db_name')  # Recibir el nombre de la base de datos del formulario
        if db_name:
            kpi_data = obtener_datos_kpi(db_name)
            clicks_data = obtener_datos_clicks(db_name)
            generar_graficos_excel(kpi_data, clicks_data)
            return send_file("kpi_clicks_report.xlsx", as_attachment=True)
        else:
            return "Por favor, selecciona una base de datos para generar los gráficos.", 400
    else:
        # Renderiza la página HTML cuando el método es GET
        return render_template('graficos.html')

# Ruta para la tercera opción, actualmente pendiente de definir
@app.route('/pendiente')
def pendiente():
    return "<h1>Esta función está pendiente de implementación</h1>"


# Configuración de la base de datos
DATABASE_URI = "mssql+pyodbc://sa:infinity@192.168.201.12/HERMESV5_HISTORICO?driver=ODBC+Driver+17+for+SQL+Server"
engine = create_engine(DATABASE_URI)

# Configurar el logging
LOG_FOLDER = r'C:\env\mi_proyecto_web\logs'  # Cambiar a ruta absoluta
if not os.path.exists(LOG_FOLDER):
    os.makedirs(LOG_FOLDER)

# Configurar el logging para crear un archivo diario
def configurar_logging():
    log_filename = os.path.join(LOG_FOLDER, f"{datetime.now().strftime('%Y-%m')}.log")
    logging.basicConfig(
        filename=log_filename,
        level=logging.INFO,
        format='%(asctime)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    print(f"Logging configurado correctamente en: {log_filename}")

configurar_logging()

# Crear la carpeta 'uploads' si no existe
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)


# Función para enviar correos
def enviar_correo(asunto, cuerpo):
    destinatario = "fdargallo@telcohumanmedia.com"
    remitente = "fdargallo@telcohumanmedia.com"
    password = "Frasco401*"

    msg = MIMEText(cuerpo)
    msg['Subject'] = asunto
    msg['From'] = remitente
    msg['To'] = destinatario

    try:
        with smtplib.SMTP('smtp.office365.com', 587) as server:
            server.starttls()
            server.login(remitente, password)
            server.send_message(msg)
        print("Correo enviado correctamente.")
    except Exception as e:
        print(f"Error al enviar el correo: {e}")

# Variable para rastrear si el informe ya fue enviado este mes
informe_enviado = False

def enviar_informe_mensual():
    global informe_enviado
    if informe_enviado:
        return  # Si ya se envió, no hacer nada

    log_filename = os.path.join(LOG_FOLDER, f"{(datetime.now() - pd.DateOffset(months=1)).strftime('%Y-%m')}.log")
    if os.path.exists(log_filename):
        enviar_correo("Informe mensual", f"Adjunto el informe del mes anterior: {log_filename}")
    else:
        enviar_correo("Informe mensual", "No había fichero disponible para el mes anterior.")
    
    informe_enviado = True  # Marcar como enviado

# Crear el scheduler
scheduler = BackgroundScheduler()

# Programar la tarea para enviar el informe mensual
scheduler.add_job(enviar_informe_mensual, 'cron', day=1, hour=0)  # Ejecutar el día 1 de cada mes a la medianoche


# Función para registrar en el log los datos específicos
def log_usuario(nombre_usuario, nombre_archivo, registros_originales, registros_limpios):
    logging.info(
        f"Usuario: {nombre_usuario}, Archivo subido: {nombre_archivo}, "
        f"Registros originales: {registros_originales}, Teléfonos limpios: {registros_limpios}"
    )
    print("Log de usuario registrado.")

def limpiar_telefono(telefono):
    if pd.isna(telefono):
        return None
    telefono = str(telefono)
    telefono = re.sub(r'[ .\-\/]', '', telefono)
    telefono = re.sub(r'^(\+34|0034)', '', telefono)
    return telefono if len(telefono) == 9 and telefono.isdigit() else None


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        flash('No file part')
        return redirect(url_for('index'))

    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        return redirect(url_for('index'))

    if file:
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)

        # Procesar el archivo Excel
        df = pd.read_excel(filepath)

        # Loguear los datos iniciales
        registros_originales = len(df)
        nombre_usuario = getpass.getuser()  # Obtener el nombre del usuario logueado

        # Renombrar las columnas a un formato estándar
        df.rename(columns=lambda x: x.strip().lower(), inplace=True)
        df.rename(columns={
            'telefono': 'telefono',
            'teléfono': 'telefono',
            'telf1': 'telefono',
            'tel1': 'telefono',
            'telefono1': 'telefono',
            'TELEFONO1': 'telefono',
            'TELEFONO': 'telefono',
            'phone': 'telefono'
        }, inplace=True)

        # Verificar si 'telefono' existe
        if 'telefono' not in df.columns:
            raise KeyError("No se encontró la columna 'telefono' en el DataFrame.")

        # Limpiar los teléfonos
        df['telefono_limpio'] = df['telefono'].apply(limpiar_telefono)
        registros_limpios = df['telefono_limpio'].notnull().sum()

        # Registrar la actividad en el log
        log_usuario(nombre_usuario, file.filename, registros_originales, registros_limpios)

        # Conectar a la base de datos y ejecutar la lógica de scoring
        with engine.connect() as conn:
            for index, row in df.iterrows():
                telefono = str(row['telefono'])

                # Consulta de scoring
                query_scoring = text("""
                    SELECT CASE 
                        WHEN COUNT(*) >= 20 THEN (CAST(SUM(CASE WHEN CallStatusNum < 11 THEN 1 ELSE 0 END) AS FLOAT) / COUNT(*)) * 100 
                        ELSE (SUM(CASE WHEN CallStatusNum < 11 THEN 1 ELSE 0 END) / 20.0) * 100 
                    END AS conteo 
                    FROM ODCalls WHERE ANI = :telefono
                """)
                result_scoring = conn.execute(query_scoring, {'telefono': telefono}).fetchone()
                conteo = result_scoring[0] if result_scoring else 0
                df.at[index, 'scoring'] = conteo

                # Consulta de media de intentos para contacto
                query_media_intentos = text("""
                    SELECT CASE 
                        WHEN COUNT(*) >= 20 THEN (COUNT(*) / CAST(NULLIF(SUM(CASE WHEN CallStatusNum < 11 THEN 1 ELSE 0 END), 0) AS FLOAT)) 
                        ELSE (20.0 / NULLIF(SUM(CASE WHEN CallStatusNum < 11 THEN 1 ELSE 0 END), 0))  
                    END AS conteo 
                    FROM ODCalls 
                    WHERE ANI = :telefono AND Duration > 1
                """)

                result_media_intentos = conn.execute(query_media_intentos, {'telefono': telefono}).fetchone()
                media_intentos = result_media_intentos[0] if result_media_intentos else None
                df.at[index, 'media_intentos_para_contacto'] = media_intentos

                # Consulta de media de intentos para contacto positivo
                query_media_intentos_positivo = text("""
                    SELECT 
                        CASE 
                            WHEN COUNT(CASE WHEN Duration > 1 AND CallStatusNum BETWEEN 1 AND 10 THEN 1 END) = 0 THEN 0
                            ELSE COUNT(*) * 1.0 / COUNT(CASE WHEN Duration > 1 AND CallStatusNum BETWEEN 1 AND 10 THEN 1 END)
                        END AS media_llamadas_para_contacto_positivo
                    FROM ODCalls
                    WHERE ANI = :telefono AND (Duration > 1 OR CallStatusNum BETWEEN 1 AND 10)
                """)

                result_media_intentos_positivo = conn.execute(query_media_intentos_positivo, {'telefono': telefono}).fetchone()
                
                
                
                # Agregar un log para depuración
                if result_media_intentos_positivo is None:
                    logging.info(f'No se encontraron resultados para el teléfono: {telefono}')
                #else:
                    #logging.info(f'Resultado para el teléfono {telefono}: {result_media_intentos_positivo[0]}')

                media_intentos_positivo = result_media_intentos_positivo[0] if result_media_intentos_positivo else None
                df.at[index, 'media_intentos_para_contacto_positivo'] = media_intentos_positivo

        # Ordenar el DataFrame de mayor a menor según el scoring
        df.sort_values(by='scoring', ascending=False, inplace=True)

        # Convertir los valores numéricos pequeños a un formato más legible con 2 decimales
        df['media_intentos_para_contacto_positivo'] = df['media_intentos_para_contacto_positivo'].apply(lambda x: '{:.2f}'.format(x) if pd.notna(x) else x)

        # Guardar el archivo Excel con el scoring, media de intentos y todos los datos
        output_filepath = os.path.join(UPLOAD_FOLDER, 'resultados_Scoring.xlsx')
        df.to_excel(output_filepath, index=False)

        # Aplicar estilo de colores en la columna 'scoring' usando openpyxl
        wb = load_workbook(output_filepath)
        ws = wb.active
        scoring_column_index = ws.max_column - 2  # Columna 'scoring'
        media_intentos_column_index = ws.max_column - 1  # Columna 'media_intentos_para_contacto'
        media_intentos_positivo_column_index = ws.max_column  # Última columna es 'media_intentos_para_contacto_positivo'

        for row in range(2, ws.max_row + 1):  # Empezamos en la fila 2 para evitar el encabezado
            score = ws.cell(row=row, column=scoring_column_index).value
            if score is not None:
                # Definir el color basado en el score (de verde a rojo)
                red = int(min(255, (255 - int(2.55 * score))))
                green = int(min(255, int(2.55 * score)))
                fill = PatternFill(start_color=f"{red:02X}{green:02X}00", end_color=f"{red:02X}{green:02X}00", fill_type="solid")
                
                ws.cell(row=row, column=scoring_column_index).fill = fill

        wb.save(output_filepath)
        wb.close()

        # Generar la tabla HTML para mostrar en el navegador
        resultados_limpieza = df.to_html(classes=['table', 'scoring'], justify='center')

        return render_template('resultados.html', tabla=resultados_limpieza)

@app.route('/download')
def download_file():
    # Descargar el archivo de resultados
    output_filepath = os.path.join(UPLOAD_FOLDER, 'resultados_Scoring.xlsx')
    return send_file(output_filepath, as_attachment=True)


if __name__ == '__main__':
    configurar_logging()
    # Prueba de envío de correo después de procesar el archivo
    enviar_correo("Prueba de correo", "Este es un mensaje de prueba para verificar el envío de correos.")
    if datetime.now().day == 1:
        enviar_informe_mensual()
    if not scheduler.running:
        scheduler.start()  # Inicia el scheduler solo si no está en ejecución
    app.run(debug=True, host='127.0.0.1', port=5000)
