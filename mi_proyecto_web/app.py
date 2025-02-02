"""
Archivo: app.py  
Descripción: Procesa archivos Excel, limpia datos y los guarda en una base de datos SQL Server.  

Autores:  
- Federico Dargallo (GitHub: Fede80)  
- Kevin Rubi (GitHub: kevoru18)  

Fecha de creación: 01/09/2024  
Última actualización: 10/01/2025  

Tecnologías utilizadas:  
- Python  
- Flask (framework web)  
- SQL Server (base de datos)  

Estructura del código:  
1. Importaciones  
2. Conexión con la base de datos  
3. Funciones para procesamiento de datos  
4. Rutas Flask  

Licencia:  
Este código está bajo la licencia MIT.
"""


from flask import Flask, render_template, request, redirect, session, url_for, flash, send_file, jsonify, send_from_directory, Blueprint,Response
import time
from flask_login import login_user, logout_user,login_required,LoginManager 
import math
from sqlalchemy import create_engine, text
from flask_sqlalchemy import SQLAlchemy 
import pandas as pd
import os
import logging
import pyodbc
import getpass  # Importar el módulo para obtener el nombre del usuario
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill 
from apscheduler.schedulers.background import BackgroundScheduler
import smtplib
from email.mime.text import MIMEText
from db_utils import obtener_datos_kpi, obtener_datos_clicks
from excel_utils import generar_graficos_excel
import re
import io
import base64
from matplotlib.figure import Figure
import plotly.graph_objs as go
import subprocess
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib
matplotlib.use('Agg')  # Establece el backend 'Agg' para evitar el error de Tcl/Tk
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
# from Model.auth import User, db
app = Flask(__name__)
# app = Blueprint('auth', __name__)
app.secret_key = "your_secret_key" #Esto es para evitar que los datos sean visibles en el navegador

# configuracion de base de datos para login

# Configuración para conectar con SQL Server
app.config['SQLALCHEMY_BINDS'] = {
# (
   'rrhh':  'mssql+pyodbc://sa:infinity@192.168.201.12/_RRHH?driver=ODBC+Driver+17+for+SQL+Server',
    'krubi_tests': 'mssql+pyodbc://sa:infinity@192.168.201.12/_krubi_tests?driver=ODBC+Driver+17+for+SQL+Server'
}
# )  # Cadena de conexión para la base de datos. Recomendación: usar variables de entorno para evitar exponer credenciales.

# esta es la buena
# # Configuración para conectar con SQL Server
# app.config['SQLALCHEMY_DATABASE_URI'] = (
#     'mssql+pyodbc://sa:infinity@192.168.201.12/_RRHH?driver=ODBC+Driver+17+for+SQL+Server'
# )

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False  # Desactiva el seguimiento de modificaciones para mejorar el rendimiento.

# Inicializamos SQLAlchemy con la configuración de Flask
db = SQLAlchemy(app)  # Vincula la instancia de la base de datos a la aplicación Flask.


# # Segunda base de datos
# app.config['SQLALCHEMY_BINDS'] = {
#     'db2': 'mssql+pyodbc://sa:password@192.168.201.12/USUARIOS_APP?driver=ODBC+Driver+17+for+SQL+Server'
# }
# db2 = SQLAlchemy()
# db2.init_app(app)




# Definimos el modelo de datos para los usuarios
class User(db.Model):  # La clase `User` hereda de `db.Model`, lo que la convierte en un modelo de SQLAlchemy.
    __bind_key__ = 'rrhh'
    __tablename__ = 'Contactos'  # Nombre de la tabla en la base de datos.
  # Mapeo de columnas
    id = db.Column(db.Integer, primary_key=True)  # Corresponde al campo `ID`.
    apellido1 = db.Column(db.String(100), nullable=False)  # Corresponde al campo `Apellido1`.
    apellido2 = db.Column(db.String(100), nullable=True)  # Corresponde al campo `Apellido2`.
    nombre = db.Column(db.String(100), nullable=False)  # Corresponde al campo `Nombre`.
    estado_telco = db.Column(db.String(50), nullable=True)  # Corresponde al campo `estado_telco`.
    categoria_telco = db.Column(db.String(50), nullable=True)  # Corresponde al campo `categoria_telco`.
    usuario_windows = db.Column(db.String(100), unique=True, nullable=False)  # Corresponde al campo `usuario_windows`.
    actualizado = db.Column(db.DateTime, nullable=True)  # Corresponde al campo `ACTUALIZADO`.
    # Método para establecer la contraseña del usuario
    def set_password(self, password):
        """
        Genera un hash seguro para la contraseña proporcionada y lo almacena en `password_hash`.
        Valida primero la complejidad de la contraseña antes de encriptarla.
        """
        self.password_hash = generate_password_hash(password, method='pbkdf2:sha256')  # Hash seguro para la contraseña.
        # if not self.validate_password(password):  # Verificamos si la contraseña cumple los criterios de complejidad.
        #     raise ValueError("Invalid password")  # Si no cumple, se lanza un error.
        # self.password_hash = generate_password_hash(password, method='pbkdf2:sha256')  # Hash seguro para la contraseña.

    # Método para verificar una contraseña
    def check_password(self, password):
        """
        Compara la contraseña proporcionada con la almacenada encriptada.
        Devuelve `True` si coinciden; de lo contrario, `False`.
        """
        return check_password_hash(self.password_hash, password)

    # Método estático para validar contraseñas
    @staticmethod
    def validate_password(password):
        """
        Verifica si una contraseña cumple con los criterios de seguridad:
        - Al menos 8 caracteres.
        - Al menos una letra mayúscula.
        - Al menos una letra minúscula.
        - Al menos un número.
        """
        return (
            len(password) >= 8 and  # Longitud mínima.
            re.search(r'[A-Z]', password) and  # Al menos una letra mayúscula.
            re.search(r'[a-z]', password) and  # Al menos una letra minúscula.
            re.search(r'\d', password)  # Al menos un dígito.
        )

    # Método estático para validar nombres de usuario
    @staticmethod
    def validate_username(username):
        """
        Verifica que el nombre de usuario solo contenga letras, números y guiones bajos.
        """
        return re.match(r'^[a-zA-Z0-9_]+$', username) is not None






class UsuariosApp(db.Model):
    __bind_key__ = 'krubi_tests'  # Indica que este modelo pertenece a la segunda base de datos
    __tablename__ = 'USUARIOS_APP'  # Asegúrate de que coincida con el nombre de la tabla en la base de datos

    id = db.Column(db.Integer, primary_key=True)  # Clave primaria
    username = db.Column(db.String(50), unique=True, nullable=False)  # Nombre de usuario único y obligatorio
    password_hash = db.Column(db.String(255), nullable=False)  # Hash de la contraseña
    email = db.Column(db.String(100), unique=True, nullable=False)  # Correo único y obligatorio
    is_active = db.Column(db.Boolean, default=True)  # Estado activo por defecto
    created_at = db.Column(db.DateTime, default=db.func.getdate())  # Fecha de creación por defecto
    last_login = db.Column(db.DateTime, nullable=True)  # Último inicio de sesión

    # def __repr__(self):
    #     return f"<UserDetails(username={self.username}, email={self.email}, is_active={self.is_active})>"









# Configurar logging
logging.basicConfig(level=logging.INFO)
# Crear la instancia de SQLAlchemy
# db = SQLAlchemy()

@app.route('/')
def inicio():
    # Esto debe ir en la parte del raiz 
    if "username" in session:
        return render_template('inicio.html',username=session['username'])
    return render_template('login.html')
    

@app.route('/index')
def index():
    if "username" in session:
        return render_template('index.html')
    return render_template('login.html')

@app.route('/main_app')
def main_app():
    return redirect(url_for('index'))  # Redirige a la función de carga de archivos

# Ruta para manejar la página de inicio de sesión
@app.route('/login', methods=['POST'])
def login():
    # Verifica si el método de la solicitud es POST (cuando se envía el formulario)
    if request.method == 'POST':
        data=request.get_json()
        # Obtiene el nombre de usuario enviado desde el formulario
        username = data.get('username')
      
         # Obtiene la contraseña enviada desde el formulario
        password = data.get('password')
        print('Username recibido:', username)
        print('Username recibido:', password)
        # Validación del lado del servidor: verifica que el nombre de usuario tenga un formato válido
        # if not User.validate_username(username):
        #     # Si el formato no es válido, devuelve un mensaje de error en formato JSON
        #     return jsonify({'success': False, 'message': 'Invalid username'})

        # Busca en la base de datos un usuario con el nombre proporcionado que esté activo
        user = User.query.filter_by(usuario_windows=username).first() #user = User.query.filter_by(username=username, is_active=True).first()
        
        # Si se encuentra el usuario y la contraseña coincide con la almacenada
        # if user and user.check_password(password):
        if user and user.estado_telco == 'ACTIVO' and user.categoria_telco == 'STAFF':
            apellido1 = user.apellido1
            nombre = user.nombre
            usuario= user.usuario_windows
            ID_usuario = user.id
            ID_usuario_texto = str(ID_usuario)
            print('Username recibido:', nombre + ' ' + apellido1+ ' ' + usuario + ' ' + ID_usuario_texto)
            # Actualiza la fecha y hora del último inicio de sesión
            # user.last_login = datetime.utcnow()
            # Guarda los cambios en la base de datos
            # db.session.commit()
            # Inicia la sesión del usuario utilizando Flask-Login
            # login_user(user)
            session['username']=username
            # clave_input=User.set_password(password,password)
            # print('Password input', clave_input)
            # # print('Entro en login')
                # 2. Buscar en UsuariosApp (prueba) usando el ID del contacto
            usuario_app = UsuariosApp.query.filter_by(id=ID_usuario).first()
            # app_username = usuario_app.username
            if check_password_hash(usuario_app.password_hash, password):
                print('Password correcta', usuario_app.password_hash)
                # Hash and update password
                # password_hash = set_password(new_password)
                # Renderiza la plantilla 'inicio.html' en caso de inicio de sesión exitoso
                # print('Entro en password')
                    # 2. Buscar en UsuariosApp (prueba) usando el ID del contacto
                # Renderiza la plantilla 'inicio.html' en caso de inicio de sesión exitoso
                # print('Usuario App recibido:', app_username)
                return jsonify({'success': True, 'redirect': '/inicio'})
            
        
            
        
        # Si el usuario no existe o la contraseña es incorrecta, devuelve un mensaje de error
        return jsonify({'success': False, 'message': 'Usuario o Contraseña erroneo'}),401 
    
    # Si el método no es POST (es GET), muestra la página de inicio de sesión
    return render_template('login.html')
# Ver esto:https://www.youtube.com/watch?v=Fr2MxT9M0V4

# Ruta para cerrar sesión
@app.route('/logout')  
# @login_required  # Decorador que asegura que esta ruta solo pueda ser accedida por usuarios autenticados.
def logout():
    # # logout_user()
    session.pop('username', None) 
    return render_template('login.html')

    # return jsonify({'success': True, 'redirect': '/login'})
   


def get_connection(db_name):
    try:
        connection = pyodbc.connect(
            f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER=192.168.201.12;DATABASE={db_name};Trusted_Connection=yes;"
        )
        logging.info("Conexión establecida correctamente con la base de datos.")
        return connection
    except pyodbc.Error as e:
        logging.error(f"Error al conectar con la base de datos: {e}")
        raise Exception(f"Error al conectar con la base de datos: {e}")

def ejecutar_consulta(connection, query):
    try:
        cursor = connection.cursor()
        cursor.execute(query)
        rows = cursor.fetchall()
        columns = [column[0] for column in cursor.description]
        logging.info(f"Consulta ejecutada correctamente: {query}")
        logging.info(f"Filas devueltas: {rows}")
        return [dict(zip(columns, row)) for row in rows]
    except pyodbc.Error as e:
        logging.error(f"Error al ejecutar la consulta: {e}")
        return f"Error al ejecutar la consulta: {e}"



def generar_graficos(resultados, db_name):
    pdf_buffer = io.BytesIO()
    imagenes_base64 = []

    logging.info(f"Generando gráficos para la base de datos: {db_name}")

    # Crear gráficos
    if resultados and resultados[0] and "TotalRecords" in resultados[0][0]:
        fig1, ax1 = plt.subplots()
        labels = ['Total Records', 'Complete Records', 'Correction Records']
        sizes = [
            resultados[0][0].get('TotalRecords', 0),
            resultados[0][0].get('CompleteRecords', 0),
            resultados[0][0].get('CorrectionRecords', 0)
        ]
        ax1.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
        ax1.axis('equal')
        plt.title('Distribución de Registros con CNAE')
        
        # Añadir leyenda
        total_registros = resultados[0][0].get('TotalRecords', 0)
        registros_con_cnae = resultados[0][0].get('CompleteRecords', 0)
        registros_sin_cnae = total_registros - registros_con_cnae
        patches = [
            mpatches.Patch(color='blue', label=f'Total de registros: {total_registros}'),
            mpatches.Patch(color='orange', label=f'Total de Registros con CNAE: {registros_con_cnae}'),
            mpatches.Patch(color='green', label=f'Total de Registros sin CNAE: {registros_sin_cnae}')
        ]
        ax1.legend(handles=patches, loc="best")
        
        buffer1 = io.BytesIO()
        fig1.savefig(buffer1, format='png')
        buffer1.seek(0)
        imagenes_base64.append(base64.b64encode(buffer1.read()).decode('utf-8'))
        plt.close(fig1)
        logging.info("Gráfico 1 generado con éxito.")
    else:
        imagenes_base64.append(None)
        logging.warning("No hay datos para el gráfico 1.")

    # Modificación del segundo gráfico:
    if resultados and resultados[1] and "ClickDate" in resultados[1][0]:
        fig2, ax2 = plt.subplots()
        fechas = [row['ClickDate'] for row in resultados[1]]
        clicks = [row['ClickCount'] for row in resultados[1]]
        
        # Añadir distinción visual
        opciones_linkedin = [clicks[i] if i % 2 == 0 else 0 for i in range(len(clicks))]
        telefono_alternativo = [clicks[i] if i % 2 != 0 else 0 for i in range(len(clicks))]

        ax2.bar(fechas, opciones_linkedin, color='blue', label='LinkedIn')
        ax2.bar(fechas, telefono_alternativo, color='green', bottom=opciones_linkedin, label='Teléfono Alternativo')

        plt.title('Clicks por Fecha')
        plt.xlabel('Fecha')
        plt.ylabel('Cantidad de Clicks')
        plt.xticks(rotation=45, ha='right')  # Rotar etiquetas de fechas para que se vean mejor
        plt.legend()
        
        buffer2 = io.BytesIO()
        fig2.savefig(buffer2, format='png')
        buffer2.seek(0)
        imagenes_base64.append(base64.b64encode(buffer2.read()).decode('utf-8'))
        plt.close(fig2)
        logging.info("Gráfico 2 generado con éxito.")
    else:
        imagenes_base64.append(None)
        logging.warning("No hay datos para el gráfico 2.")

    if resultados and resultados[2] and "Total" in resultados[2][0]:
        fig3, ax3 = plt.subplots()
        labels = ['Total', 'Contactadas']
        values = [
            resultados[2][0].get('Total', 0),
            resultados[2][0].get('Contactadas', 0)
        ]
        ax3.bar(labels, values, color=['green', 'orange'])
        plt.title('Contactadas vs Total')
        buffer3 = io.BytesIO()
        fig3.savefig(buffer3, format='png')
        buffer3.seek(0)
        imagenes_base64.append(base64.b64encode(buffer3.read()).decode('utf-8'))
        plt.close(fig3)
        logging.info("Gráfico 3 generado con éxito.")
    else:
        imagenes_base64.append(None)
        logging.warning("No hay datos para el gráfico 3.")

    # Guardar gráficos en PDF
    with PdfPages(pdf_buffer) as pdf:
        for i, img in enumerate(imagenes_base64):
            if img:
                fig, ax = plt.subplots()
                ax.imshow(plt.imread(io.BytesIO(base64.b64decode(img))))
                ax.axis('off')
                pdf.savefig(fig)
                plt.close(fig)

    pdf_buffer.seek(0)
    return imagenes_base64, pdf_buffer

@app.route('/graficos', methods=['GET', 'POST'])
def graficos():
    if "username" in session:
        if request.method == 'POST':
            db_name = request.form.get('db_name')

            if not db_name:
                return "Por favor, introduce una base de datos.", 400

            try:
                connection = get_connection(db_name)

                query1 = "SELECT TotalRecords, CompleteRecords, CorrectionRecords FROM CNAE_KPI_Audit"
                query2 = "SELECT CAST(ClickDate AS DATE) AS ClickDate, COUNT(*) AS ClickCount FROM LinkedinClickLog GROUP BY CAST(ClickDate AS DATE)"
                query3 = "SELECT COUNT(*) AS Total, SUM(CASE WHEN PERSONA_CONTACTADA IS NOT NULL THEN 1 ELSE 0 END) AS Contactadas FROM Empresas"

                resultados = [
                    ejecutar_consulta(connection, query1),
                    ejecutar_consulta(connection, query2),
                    ejecutar_consulta(connection, query3)
                ]

            # Log de los resultados de las consultas
                logging.info(f"Resultados de la consulta 1: {resultados[0]}")
                logging.info(f"Resultados de la consulta 2: {resultados[1]}")
                logging.info(f"Resultados de la consulta 3: {resultados[2]}")

                imagenes_base64, pdf_buffer = generar_graficos(resultados, db_name)

                return render_template('graficos.html', imagenes_base64=imagenes_base64, pdf_link=url_for('descargar_pdf_file', db_name=db_name))
            except Exception as e:
                logging.error(f"Ocurrió un error: {e}")
                return f"Ocurrió un error: {e}", 500

        return render_template('graficos.html')
    return render_template('login.html')
@app.route('/descargar_pdf/<db_name>', methods=['GET'])
def descargar_pdf_file(db_name):
    try:
        connection = get_connection(db_name)

        query1 = "SELECT TotalRecords, CompleteRecords, CorrectionRecords FROM CNAE_KPI_Audit"
        query2 = "SELECT CAST(ClickDate AS DATE) AS ClickDate, COUNT(*) AS ClickCount FROM LinkedinClickLog GROUP BY CAST(ClickDate AS DATE)"
        query3 = "SELECT COUNT(*) AS Total, SUM(CASE WHEN PERSONA_CONTACTADA IS NOT NULL THEN 1 ELSE 0 END) AS Contactadas FROM Empresas"

        resultados = [
            ejecutar_consulta(connection, query1),
            ejecutar_consulta(connection, query2),
            ejecutar_consulta(connection, query3)
        ]

        _, pdf_buffer = generar_graficos(resultados, db_name)

        return send_file(pdf_buffer, as_attachment=True, download_name=f"graficos_{db_name}_{datetime.now().strftime('%Y%m%d')}.pdf", mimetype='application/pdf')
    except Exception as e:
        logging.error(f"Ocurrió un error: {e}")
        return f"Ocurrió un error: {e}", 500


@app.route('/descargar-excel')
def descargar_excel():
    excel_path = os.path.join(os.getcwd(), "env", "mi_proyecto_web", "mi_proyecto_web", "kpi_clicks_report.xlsx")
    return send_file(excel_path, as_attachment=True)

# Ruta para la tercera opción, actualmente pendiente de definir
@app.route('/pendiente')
def pendiente():
    return "<h1>Esta función está pendiente de implementación</h1>"


# Configuración de la base de datos
DATABASE_URI = "mssql+pyodbc://sa:infinity@192.168.201.12/HERMESV5_HISTORICO?driver=ODBC+Driver+17+for+SQL+Server"
engine = create_engine(DATABASE_URI)

# Configurar el logging
LOG_FOLDER = r'C:\env\mi_proyecto_web\logs'  # Cambiar a ruta absoluta
if not os.path.exists(LOG_FOLDER):
    os.makedirs(LOG_FOLDER)

# Configurar el logging para crear un archivo diario
def configurar_logging():
    log_filename = os.path.join(LOG_FOLDER, f"{datetime.now().strftime('%Y-%m')}.log")
    logging.basicConfig(
        filename=log_filename,
        level=logging.INFO,
        format='%(asctime)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    print(f"Logging configurado correctamente en: {log_filename}")

configurar_logging()

# Crear la carpeta 'uploads' si no existe
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)


# Función para enviar correos
def enviar_correo(asunto, cuerpo):
    destinatario = "fdargallo@telcohumanmedia.com"
    remitente = "fdargallo@telcohumanmedia.com"
    password = "Frasco401*"

    msg = MIMEText(cuerpo)
    msg['Subject'] = asunto
    msg['From'] = remitente
    msg['To'] = destinatario

    try:
        with smtplib.SMTP('smtp.office365.com', 587) as server:
            server.starttls()
            server.login(remitente, password)
            server.send_message(msg)
        print("Correo enviado correctamente.")
    except Exception as e:
        print(f"Error al enviar el correo: {e}")

# Variable para rastrear si el informe ya fue enviado este mes
informe_enviado = False

def enviar_informe_mensual():
    global informe_enviado
    if informe_enviado:
        return  # Si ya se envió, no hacer nada

    log_filename = os.path.join(LOG_FOLDER, f"{(datetime.now() - pd.DateOffset(months=1)).strftime('%Y-%m')}.log")
    if os.path.exists(log_filename):
        enviar_correo("Informe mensual", f"Adjunto el informe del mes anterior: {log_filename}")
    else:
        enviar_correo("Informe mensual", "No había fichero disponible para el mes anterior.")
    
    informe_enviado = True  # Marcar como enviado

# Crear el scheduler
scheduler = BackgroundScheduler()

# Programar la tarea para enviar el informe mensual
scheduler.add_job(enviar_informe_mensual, 'cron', day=1, hour=0)  # Ejecutar el día 1 de cada mes a la medianoche


# Función para registrar en el log los datos específicos
def log_usuario(nombre_usuario, nombre_archivo, registros_originales, registros_limpios):
    logging.info(
        f"Usuario: {nombre_usuario}, Archivo subido: {nombre_archivo}, "
        f"Registros originales: {registros_originales}, Teléfonos limpios: {registros_limpios}"
    )
    print("Log de usuario registrado.")
    
    # Enviar un evento SSE al cliente
    def generate_event():
        yield "data: iniciarCarga()\n\n"  # Indica al cliente que ejecute la función
    return Response(generate_event(), mimetype='text/event-stream')

def limpiar_telefono(telefono):
    if pd.isna(telefono):
        return None
    telefono = str(telefono)
    telefono = re.sub(r'[ .\-\/]', '', telefono)
    telefono = re.sub(r'^(\+34|0034)', '', telefono)
    return telefono if len(telefono) == 9 and telefono.isdigit() else None

# progress = 0
@app.route('/upload', methods=['POST'])
def upload_file():
    if "username" not in session:
        return render_template('login.html')
    # global progress
    # # nuevo
    # progress = 0  # Reiniciar progreso
    # total_steps = 100

    if 'file' not in request.files or request.files['file'].filename == '':
        flash('Archivo no válido')
        return redirect(url_for('index'))


    file = request.files['file']
    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)
    # if file:
    try:
        # Procesar el archivo Excel
        df = pd.read_excel(filepath)
        # Loguear los datos iniciales
        registros_originales = len(df)
        nombre_usuario = getpass.getuser()  # Obtener el nombre del usuario logueado
        # Renombrar las columnas a un formato estándar
        column_mapping={
            'telefono': 'telefono',
            'teléfono': 'telefono',
            'telf1': 'telefono',
            'tel1': 'telefono',
            'telefono1': 'telefono',
            'TELEFONO1': 'telefono',
            'TELEFONO': 'telefono',
            'phone': 'telefono'
        }
        df.rename(columns=lambda x: x.strip().lower(), inplace=True)
        df.rename(columns=column_mapping, inplace=True)

    # Verificar si 'telefono' existe
        if 'telefono' not in df.columns:
            raise KeyError("No se encontró la columna 'telefono' en el archivo.")

    # Limpiar los teléfonos
        # total_steps = len(df) + 3  # Procesos principales + registros
        df['telefono_limpio'] = df['telefono'].apply(limpiar_telefono)
        registros_limpios = df['telefono_limpio'].notnull().sum()
  # Registrar la actividad en el log
        log_usuario(nombre_usuario, file.filename, registros_originales, registros_limpios)

    # Conectar a la base de datos y ejecutar la lógica de scoring
        with engine.connect() as conn:
            # total_steps = 3  # Número total de pasos
            # current_step = 1
        # send_progress_update(current_step, total_steps)
            for index, row in df.iterrows():
                telefono = str(row['telefono'])

            # Consulta de scoring
                query_scoring = text("""
                    SELECT CASE 
                        WHEN COUNT(*) >= 20 THEN (CAST(SUM(CASE WHEN CallStatusNum < 11 THEN 1 ELSE 0 END) AS FLOAT) / COUNT(*)) * 100 
                        ELSE (SUM(CASE WHEN CallStatusNum < 11 THEN 1 ELSE 0 END) / 20.0) * 100 
                    END AS conteo 
                    FROM ODCalls WHERE ANI = :telefono
                """)
                result_scoring = conn.execute(query_scoring, {'telefono': telefono}).fetchone()
                # nuevo comentario
                # conteo = result_scoring[0] if result_scoring else 0
                # df.at[index, 'scoring'] = conteo
                
                df.at[index, 'scoring'] = result_scoring[0] if result_scoring else 0

            # Consulta de media de intentos para contacto
                query_media_intentos = text("""
                    SELECT CASE 
                        WHEN COUNT(*) >= 20 THEN (COUNT(*) / CAST(NULLIF(SUM(CASE WHEN CallStatusNum < 11 THEN 1 ELSE 0 END), 0) AS FLOAT)) 
                        ELSE (20.0 / NULLIF(SUM(CASE WHEN CallStatusNum < 11 THEN 1 ELSE 0 END), 0))  
                    END AS conteo 
                    FROM ODCalls 
                    WHERE ANI = :telefono AND Duration > 1
                """)

                result_media_intentos = conn.execute(query_media_intentos, {'telefono': telefono}).fetchone()
                df.at[index, 'media_intentos_para_contacto'] = result_media_intentos[0] if result_media_intentos else None

            # Consulta de media de intentos para contacto positivo
                query_media_intentos_positivo = text("""
                    SELECT 
                        CASE 
                            WHEN COUNT(CASE WHEN Duration > 1 AND CallStatusNum BETWEEN 1 AND 10 THEN 1 END) = 0 THEN 0
                            ELSE COUNT(*) * 1.0 / COUNT(CASE WHEN Duration > 1 AND CallStatusNum BETWEEN 1 AND 10 THEN 1 END)
                        END AS media_llamadas_para_contacto_positivo
                    FROM ODCalls
                    WHERE ANI = :telefono AND (Duration > 1 OR CallStatusNum BETWEEN 1 AND 10)
                """)

                result_media_intentos_positivo = conn.execute(query_media_intentos_positivo, {'telefono': telefono}).fetchone()
                # Agregar un log para depuración
                if result_media_intentos_positivo is None:
                    logging.info(f'No se encontraron resultados para el teléfono: {telefono}')
      
                media_intentos_positivo = result_media_intentos_positivo[0] if result_media_intentos_positivo else None
                df.at[index, 'media_intentos_para_contacto_positivo'] = media_intentos_positivo

    # Ordenar el DataFrame de mayor a menor según el scoring
        df.sort_values(by='scoring', ascending=False, inplace=True)
    # Limitar a los primeros 5 registros
        df = df.head(registros_originales)
        page = request.args.get("page", 1, type=int)
        
        
        
    # # queda pendiente paginacion
    #     rows_per_page = 5
    #     total_pages = math.ceil(len(data) / rows_per_page)
    #     start = (page - 1) * rows_per_page
    #     end = start + rows_per_page
    #     page_data = data[start:end]






    # Convertir los valores numéricos pequeños a un formato más legible con 2 decimales
        df['media_intentos_para_contacto_positivo'] = df['media_intentos_para_contacto_positivo'].apply(lambda x: '{:.2f}'.format(x) if pd.notna(x) else x)

    # Guardar el archivo Excel con el scoring, media de intentos y todos los datos
        output_filepath = os.path.join(UPLOAD_FOLDER, 'resultados_Scoring.xlsx')
        df.to_excel(output_filepath, index=False)

    # Aplicar estilo de colores en la columna 'scoring' usando openpyxl
        wb = load_workbook(output_filepath)
        ws = wb.active
        scoring_column_index = ws.max_column - 2  # Columna 'scoring'
        # media_intentos_column_index = ws.max_column - 1  # Columna 'media_intentos_para_contacto'
        # media_intentos_positivo_column_index = ws.max_column  # Última columna es 'media_intentos_para_contacto_positivo'

        for row in range(2, ws.max_row + 1):  # Empezamos en la fila 2 para evitar el encabezado
            score = ws.cell(row=row, column=scoring_column_index).value
            if score is not None:
            # Definir el color basado en el score (de verde a rojo)
                red = int(min(255, (255 - int(2.55 * score))))
                green = int(min(255, int(2.55 * score)))
                fill = PatternFill(start_color=f"{red:02X}{green:02X}00", end_color=f"{red:02X}{green:02X}00", fill_type="solid")
            
                ws.cell(row=row, column=scoring_column_index).fill = fill

        wb.save(output_filepath)
        wb.close()

    # Generar la tabla HTML para mostrar en el navegador
        resultados_limpieza = df.to_html(classes=['table', 'scoring'], justify='center')
        return render_template('resultados.html', tabla=resultados_limpieza) #jsonify({'progress': 100, 'results': resultados_limpieza})             #      #, 

    
    except Exception as e:
        logging.error(f"Error processing file: {e}")
        flash('Error processing file')
        return redirect(url_for('index'))

    
    
    
    
    
    
    
    # barra
@app.route('/progress', methods=['GET'])
def get_progress():
    global progress
    return jsonify({'progress': progress})





@app.route('/download')
def download_file():
    # Descargar el archivo de resultados
    output_filepath = os.path.join(UPLOAD_FOLDER, 'resultados_Scoring.xlsx')
    return send_file(output_filepath, as_attachment=True)


if __name__ == '__main__':
    configurar_logging()
    # Prueba de envío de correo después de procesar el archivo
    enviar_correo("Prueba de correo", "Este es un mensaje de prueba para verificar el envío de correos.")
    if datetime.now().day == 1:
        enviar_informe_mensual()
    if not scheduler.running:
        scheduler.start()  # Inicia el scheduler solo si no está en ejecución
    app.run(debug=True, host='127.0.0.1', port=5000)
